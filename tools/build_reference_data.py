#!/usr/bin/env python3
"""Build reference data: SWQL functions, status codes, and NetObject types.

Two inputs are merged:

1. ``docs/swql-functions/index.md`` from the OrionSDK gh-pages branch — the authoritative
   list of functions SWIS supports, with signatures and semantics.
2. A community SWQL examples workbook (``SWIS Examples`` / ``SWIS Table Reference info`` /
   ``Status IDs`` sheets) — worked examples, observed results, and the NetObjectType
   prefix table, none of which appear in the official reference.

The two disagree in places. Rather than silently picking a winner, the merge records
both and emits a ``reconciliation`` report so the discrepancies are visible and can be
verified against a live server.

Usage:
    python tools/build_reference_data.py \
        --functions-md /path/to/gh-pages/docs/swql-functions/index.md \
        --workbook /path/to/SWQL_Examples.xlsx

Outputs (under data/reference/):
    swql-functions.json    official signatures joined to worked examples
    status-codes.json      status id -> name, rank, meaning
    netobject-types.json   entity -> NetObjectType prefix, key properties, parent entity
    reconciliation.json    documented vs. observed differences
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


# --------------------------------------------------------------------------------------
# Official function reference (markdown)
# --------------------------------------------------------------------------------------

# Bullets look like: * `Name(args)` - Description text.
BULLET_RE = re.compile(r"^\*\s+`(?P<sig>[^`]+)`\s*(?:-\s*(?P<desc>.*))?$")
HEADING_RE = re.compile(r"^##\s+(?P<title>.+?)\s*$")
# "_Available in Orion Platform 2016.1 and later._" / "Requires Orion 2018.3 or later." /
# "Supported since Orion Platform 2017.3 (NPM 12.2) and later."
SINCE_RE = re.compile(
    r"(?:available in|requires|supported since)\s+(?:orion\s+)?(?:platform\s+)?"
    r"(?P<version>\d{4}\.\d+)",
    re.I,
)


def parse_functions_md(path: str) -> list[dict]:
    functions: list[dict] = []
    category = "General"
    for line in open(path, encoding="utf-8"):
        line = line.rstrip("\n")
        h = HEADING_RE.match(line)
        if h:
            category = re.sub(r"\s*Functions?$", "", h.group("title")).strip() or "General"
            continue
        m = BULLET_RE.match(line.strip())
        if not m:
            continue
        sig = m.group("sig").strip()
        desc = (m.group("desc") or "").strip()

        name = re.split(r"[(\s]", sig, 1)[0].strip()
        args_m = re.search(r"\((.*)\)", sig, re.S)
        args = args_m.group(1).strip() if args_m else ""

        entry = {
            "name": name,
            "category": category,
            "signature": sig,
            "arguments": args,
            # Some entries are documented without parentheses (e.g. `Minute`, `ToUtc`,
            # `DateTime`); they still take arguments, so flag rather than assert arity.
            "signatureComplete": bool(args_m),
            "description": re.sub(r"[_*`]", "", desc).strip(),
            "source": "OrionSDK docs/swql-functions",
        }
        since = SINCE_RE.search(desc)
        if since:
            entry["availableSince"] = since.group("version")
        functions.append(entry)
    return functions


# --------------------------------------------------------------------------------------
# Community workbook
# --------------------------------------------------------------------------------------


def cell(row, i):
    if i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v).strip()


def read_sheet(wb, name):
    if name not in wb.sheetnames:
        return []
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def parse_workbook(path: str) -> dict:
    if openpyxl is None:
        sys.exit("error: openpyxl is required to read the workbook (pip install openpyxl)")
    wb = openpyxl.load_workbook(path, data_only=True)

    examples = []
    for row in read_sheet(wb, "SWIS Examples")[1:]:
        name = cell(row, 2)
        query = cell(row, 3)
        if not name or not query:
            continue
        entry = {
            "name": name,
            "group": cell(row, 1),
            "query": " ".join(query.split()),
            "minCoreVersion": cell(row, 0),
        }
        if cell(row, 4):
            entry["observedResult"] = cell(row, 4)
        if cell(row, 5):
            entry["notes"] = " ".join(cell(row, 5).split())
        examples.append(entry)

    # One entity can occupy several rows, each naming a different key property.
    entities: dict[str, dict] = {}
    for row in read_sheet(wb, "SWIS Table Reference info")[1:]:
        entity = cell(row, 1)
        if not entity:
            continue
        rec = entities.setdefault(
            entity,
            {
                "entity": entity,
                "module": cell(row, 0),
                "displayName": cell(row, 2),
                "netObjectPrefix": cell(row, 3),
                "keyProperties": [],
                "parentEntities": [],
                "captionColumn": cell(row, 6) or None,
            },
        )
        for key in (k.strip() for k in cell(row, 4).split(",")):
            if key and key not in rec["keyProperties"]:
                rec["keyProperties"].append(key)
        parent = cell(row, 5)
        if parent and parent not in rec["parentEntities"]:
            rec["parentEntities"].append(parent)
        if not rec["netObjectPrefix"] and cell(row, 3):
            rec["netObjectPrefix"] = cell(row, 3)
        if not rec["captionColumn"] and cell(row, 6):
            rec["captionColumn"] = cell(row, 6)

    statuses = []
    for row in read_sheet(wb, "Status IDs")[1:]:
        sid = cell(row, 0)
        if not sid.isdigit():
            continue
        statuses.append(
            {
                "status": int(sid),
                "name": cell(row, 1),
                "rank": int(cell(row, 2)) if cell(row, 2).isdigit() else None,
                "description": " ".join(cell(row, 3).split()) or None,
            }
        )

    return {
        "examples": examples,
        "entities": sorted(entities.values(), key=lambda e: e["entity"]),
        "statuses": sorted(statuses, key=lambda s: s["status"]),
    }


# --------------------------------------------------------------------------------------
# Merge and reconcile
# --------------------------------------------------------------------------------------


def norm_name(name: str) -> str:
    """Match keys on a squashed name: the workbook writes 'Get Date' for GetDate."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def merge(functions: list[dict], examples: list[dict]) -> tuple[list[dict], list[dict]]:
    by_name = {norm_name(f["name"]): f for f in functions}
    notes: list[dict] = []

    for ex in examples:
        target = by_name.get(norm_name(ex["name"]))
        if target is None:
            # Present in the workbook but absent from the official reference. Keep it,
            # flagged, because undocumented-but-working functions are still useful.
            entry = {
                "name": ex["name"],
                "category": ex["group"] or "Uncategorized",
                "signature": ex["name"],
                "arguments": "",
                "signatureComplete": False,
                "description": ex.get("notes", ""),
                "source": "community workbook",
                "documented": False,
                "examples": [],
            }
            functions.append(entry)
            by_name[norm_name(ex["name"])] = entry
            notes.append(
                {
                    "type": "undocumented-function",
                    "function": ex["name"],
                    "detail": (
                        "Used in the community workbook but absent from the official "
                        "SWQL function reference. Verify against your platform version "
                        "before relying on it."
                    ),
                    "example": ex["query"],
                }
            )
            target = entry

        target.setdefault("examples", []).append(
            {k: v for k, v in ex.items() if k in ("query", "observedResult", "notes", "minCoreVersion")}
        )

        wb_ver = ex.get("minCoreVersion")
        doc_ver = target.get("availableSince")
        if wb_ver and doc_ver and wb_ver != doc_ver:
            notes.append(
                {
                    "type": "version-mismatch",
                    "function": target["name"],
                    "detail": (
                        f"Official reference says available since {doc_ver}; the workbook "
                        f"records a minimum core version of {wb_ver}."
                    ),
                    "documented": doc_ver,
                    "workbook": wb_ver,
                }
            )
        elif wb_ver and not doc_ver:
            target["workbookMinCoreVersion"] = wb_ver

    for f in functions:
        f.setdefault("documented", f["source"].startswith("OrionSDK"))
        f.setdefault("examples", [])
        f["exampleCount"] = len(f["examples"])

    functions.sort(key=lambda f: (f["category"], f["name"].lower()))
    return functions, notes


def suggest_rename(missing: str, known: set[str]) -> list[str]:
    """Propose current entity names for one that is no longer published.

    Three cheap signals cover the renames actually seen between the workbook's era and
    the current schema: a pure case change (Orion.VIM.LUNs -> Orion.VIM.Luns), a typo
    (Orion.SRM.FIleServerIdentification -> ...FileServerIdentification), and a namespace
    move that keeps the leaf (Orion.NPM.UCSBlades -> Orion.UCS.Blades).
    """
    squash = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())
    singular = lambda s: s[:-1] if len(s) > 3 and s.endswith("s") else s

    # Sorting every candidate list, and breaking ties on the name rather than leaving them
    # to set iteration order, keeps the output byte-identical between runs. Python
    # randomizes string hashing per process, so an unsorted set walk is not reproducible.
    exact = sorted(k for k in known if squash(k) == squash(missing))
    if exact:
        return exact[:1]

    leaf = squash(missing.split(".")[-1])
    leaf_sing = singular(leaf)
    missing_parts = [p.lower() for p in missing.split(".")]

    def shared_depth(candidate: str) -> int:
        n = 0
        for a, b in zip(missing_parts, [p.lower() for p in candidate.split(".")]):
            if a != b:
                break
            n += 1
        return n

    hits = []
    for k in known:
        # Same leaf, ignoring plural: Orion.F5.Pools -> Orion.F5.LTM.Pool.
        leaf_match = singular(squash(k.split(".")[-1])) == leaf_sing
        # Namespace folded into the leaf: Orion.NPM.UCSBlades -> Orion.UCS.Blades.
        folded = squash(k).endswith(leaf)
        if leaf_match or folded:
            hits.append((shared_depth(k), folded, k))

    if not hits:
        return []
    best = max(d for d, _, _ in hits)
    # Two shared namespace segments is a confident rename. One segment only counts when
    # the leaf itself is long and distinctive, so "Orion.F5.Nodes" does not resolve to
    # the unrelated "Orion.Nodes".
    if best >= 2:
        winners = [k for d, _, k in hits if d == best]
    elif best == 1 and len(leaf) >= 7:
        winners = [k for d, folded, k in hits if d == best and folded]
    else:
        return []
    return sorted(winners, key=lambda k: (len(k), k))[:3]


def cross_check_entities(entities: list[dict], schema_index_path: str) -> list[dict]:
    """Flag workbook entities that no longer exist in the current published schema."""
    if not os.path.isfile(schema_index_path):
        return []
    known = {e["entity"] for e in json.load(open(schema_index_path, encoding="utf-8"))}
    notes = []
    for rec in entities:
        rec["inCurrentSchema"] = rec["entity"] in known
        if rec["inCurrentSchema"]:
            continue
        candidates = suggest_rename(rec["entity"], known)
        if candidates:
            rec["supersededBy"] = candidates[0]
        notes.append(
            {
                "type": "unknown-entity",
                "entity": rec["entity"],
                "detail": (
                    "Listed in the workbook but not present in the published schema for "
                    "the built version. It may be renamed, module-gated, or removed."
                ),
                "likelyReplacements": candidates,
            }
        )
    return notes


def write_json(path: str, payload) -> None:
    """Write atomically, so a concurrent reader never sees a half-written file.

    Same reasoning as tools/build_schema_data.py: anything reading these while a rebuild
    runs gets a JSONDecodeError partway through, and the rename makes the swap atomic.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1, ensure_ascii=False)
        fh.write("\n")
    os.replace(tmp, path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--functions-md", required=True)
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--schema-index", default="data/schema/2026.2/index.json")
    ap.add_argument("--out", default="data/reference")
    args = ap.parse_args()

    functions = parse_functions_md(args.functions_md)
    workbook = parse_workbook(args.workbook)
    functions, notes = merge(functions, workbook["examples"])
    notes += cross_check_entities(workbook["entities"], args.schema_index)

    write_json(os.path.join(args.out, "swql-functions.json"), functions)
    write_json(os.path.join(args.out, "status-codes.json"), workbook["statuses"])
    write_json(os.path.join(args.out, "netobject-types.json"), workbook["entities"])
    write_json(os.path.join(args.out, "reconciliation.json"), notes)

    print(
        json.dumps(
            {
                "functions": len(functions),
                "documented": sum(1 for f in functions if f["documented"]),
                "withExamples": sum(1 for f in functions if f["exampleCount"]),
                "statusCodes": len(workbook["statuses"]),
                "netObjectTypes": len(workbook["entities"]),
                "reconciliationNotes": len(notes),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
